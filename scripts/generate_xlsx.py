#!/usr/bin/env python3
"""Generate KIZ0 Training xlsx that matches the original spreadsheet design."""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── Helpers ──────────────────────────────────────────────────────────────────

def rgb(r, g=None, b=None):
    if g is None:
        return f"FF{r[1:]}" if r.startswith('#') else f"FF{r}"
    return f"FF{r:02X}{g:02X}{b:02X}"

BLACK    = rgb("000000")
WHITE    = rgb("FFFFFF")
DGRAY    = rgb("0D0D0D")
GRAY1    = rgb("141414")
GRAY2    = rgb("1E1E1E")
GRAY3    = rgb("2E2E2E")
LGRAY    = rgb("EFEFEF")
LGRAY2   = rgb("F3F3F3")
RED_A    = rgb("FF3333")

def fill(color): return PatternFill("solid", fgColor=color)
def font(bold=False, color=WHITE, sz=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=sz, italic=italic)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)
def right():  return Alignment(horizontal='right',  vertical='center', wrap_text=True)
def border_all(color=GRAY3):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def border_bottom(color=GRAY3):
    s = Side(style='thin', color=color)
    return Border(bottom=s)
def border_top_bottom(color=GRAY3):
    s = Side(style='thin', color=color)
    return Border(top=s, bottom=s)

def style(ws, coord, value=None, bold=False, fc=WHITE, bg=None,
          align=None, sz=10, border=None, italic=False, num_format=None):
    c = ws[coord]
    if value is not None: c.value = value
    c.font = font(bold=bold, color=fc, sz=sz, italic=italic)
    if bg: c.fill = fill(bg)
    if align: c.alignment = align
    if border: c.border = border
    if num_format: c.number_format = num_format
    return c

def resolve_load(val, prev=None):
    """Resolve formula strings to numbers."""
    if val is None or val == '': return 0
    try: return float(val)
    except: pass
    s = str(val).strip()
    import re
    # 90%prev
    m = re.match(r'^(\d+(?:\.\d+)?)\s*%', s)
    if m and prev: return round(float(m.group(1)) / 100 * prev * 2) / 2
    # 0.9*prev
    m = re.match(r'^(\d+(?:\.\d+)?)\s*\*?\s*prev', s, re.I)
    if m and prev: return round(float(m.group(1)) * prev * 2) / 2
    # prev-10
    m = re.match(r'^prev\s*-\s*(\d+(?:\.\d+)?)', s, re.I)
    if m and prev: return prev - float(m.group(1))
    m = re.match(r'^prev\s*\+\s*(\d+(?:\.\d+)?)', s, re.I)
    if m and prev: return prev + float(m.group(1))
    return 0

def calc_1rm(load, reps):
    if reps <= 0: return 0
    if reps == 1: return load
    return round(load * (1 + reps / 30))

# ─── Main builder ─────────────────────────────────────────────────────────────

def build(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "List 15"
    ws.sheet_view.showGridLines = False

    weeks = data.get('weeks', [])
    lifter = data.get('lifterName', 'KIZ0')

    # ── Column widths ──────────────────────────────────────────────────────────
    # We'll use columns per week block: each week = 12 cols
    # Left sidebar: A=col1 (exercise), B=col2 (priority)
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18

    # Fixed left section header
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 18

    # ── Top header ────────────────────────────────────────────────────────────
    ws.merge_cells('A1:B3')
    c = ws['A1']
    c.value = lifter
    c.font = Font(name="Arial", bold=True, color=WHITE, size=22)
    c.fill = fill(BLACK)
    c.alignment = center()

    ws.merge_cells('A4:B4')
    style(ws, 'A4', 'POWERLIFTING TRACKER', bold=True, fc=WHITE, bg=BLACK, align=center(), sz=9)

    # Separator
    ws.row_dimensions[5].height = 6
    ws.merge_cells('A5:B5')
    ws['A5'].fill = fill(BLACK)

    # ── Per-week columns ──────────────────────────────────────────────────────
    # Week blocks start at column C (col 3), each block = 11 cols wide
    # Layout per week: date | sets | reps | rpe | range | notes | load | athlete_rpe | tonnage | video | (gap)
    WEEK_COLS = 11
    COL_LABELS = ['SETS', 'REPS', 'RPE', 'RANGE', 'NOTES', 'LOAD', 'RPE', 'TONNAGE', 'VIDEO']
    COL_WIDTHS = [6, 6, 6, 10, 16, 8, 6, 10, 8, 1]

    def week_start_col(w): return 3 + w * WEEK_COLS  # 1-based

    for wi, week in enumerate(weeks):
        wc = week_start_col(wi)
        wc_letter = get_column_letter(wc)

        # Set column widths for this week
        for j, cw in enumerate(COL_WIDTHS):
            ws.column_dimensions[get_column_letter(wc + j)].width = cw

        # Row 1-2: WEEK header
        ws.merge_cells(f'{get_column_letter(wc)}1:{get_column_letter(wc+9)}2')
        c = ws[f'{get_column_letter(wc)}1']
        c.value = f'WEEK {week.get("weekNumber", wi+1)}'
        c.font = Font(name="Arial", bold=True, color=WHITE, size=13)
        c.fill = fill(BLACK)
        c.alignment = center()

        # Row 3: microcycle header
        ws.merge_cells(f'{get_column_letter(wc)}3:{get_column_letter(wc+9)}3')
        c = ws[f'{get_column_letter(wc)}3']
        c.value = 'Microcycle Metrics'
        c.font = font(bold=True, color=WHITE, sz=8)
        c.fill = fill(GRAY2)
        c.alignment = center()

        # Row 4: column label row
        ws.row_dimensions[4].height = 14
        for j, lbl in enumerate(COL_LABELS):
            coord = f'{get_column_letter(wc+j)}4'
            style(ws, coord, lbl, bold=True, fc=WHITE, bg=GRAY2, align=center(), sz=7, border=border_all(BLACK))

        # Row 5: gap
        for j in range(10):
            ws[f'{get_column_letter(wc+j)}5'].fill = fill(BLACK)

    # ── Days ──────────────────────────────────────────────────────────────────
    # Find all days across all weeks to determine global row layout
    # We'll iterate days in week 0 as the "template" for rows
    # then fill other weeks in parallel columns

    if not weeks: 
        wb.save(sys.argv[2])
        return

    # Collect all days from week 0 as row blueprint
    base_week = weeks[0]
    base_days = base_week.get('days', [])

    current_row = 7  # start row for first day

    for di, day in enumerate(base_days):
        day_blocks = day.get('blocks', [])
        n_exercise_rows = sum(len(b.get('sets', [])) for b in day_blocks) if day_blocks else 3
        n_rows = max(n_exercise_rows + 3, 5)  # header + col labels + exercises + gap

        day_label = f'DAY {day.get("dayNumber", di+1)}'

        # ── Day header row ────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 18
        ws.merge_cells(f'A{current_row}:B{current_row}')
        c = ws[f'A{current_row}']
        c.value = day_label
        c.font = Font(name="Arial", bold=True, color=WHITE, size=12)
        c.fill = fill(BLACK)
        c.alignment = left()
        # Borders
        c.border = border_all(GRAY3)

        for wi in range(len(weeks)):
            wc = week_start_col(wi)
            coord_end = get_column_letter(wc + 9)
            ws.merge_cells(f'{get_column_letter(wc)}{current_row}:{coord_end}{current_row}')
            c2 = ws[f'{get_column_letter(wc)}{current_row}']
            c2.value = weeks[wi].get('days', [{}])[di].get('date', '') if di < len(weeks[wi].get('days', [])) else ''
            c2.font = font(bold=True, color=WHITE, sz=10)
            c2.fill = fill(BLACK)
            c2.alignment = center()

        # ── Coach/Athlete subheader ───────────────────────────────────────────
        current_row += 1
        ws.row_dimensions[current_row].height = 13
        ws.merge_cells(f'A{current_row}:B{current_row}')
        style(ws, f'A{current_row}', 'Exercise / Priority', bold=True, fc=WHITE, bg=GRAY2, align=left(), sz=8, border=border_all(BLACK))

        for wi in range(len(weeks)):
            wc = week_start_col(wi)
            # Coach label
            ws.merge_cells(f'{get_column_letter(wc)}{current_row}:{get_column_letter(wc+4)}{current_row}')
            c3 = ws[f'{get_column_letter(wc)}{current_row}']
            c3.value = 'Coach'
            c3.font = font(bold=True, color=WHITE, sz=8)
            c3.fill = fill(GRAY2)
            c3.alignment = center()
            # Athlete label
            ws.merge_cells(f'{get_column_letter(wc+5)}{current_row}:{get_column_letter(wc+8)}{current_row}')
            c4 = ws[f'{get_column_letter(wc+5)}{current_row}']
            c4.value = 'Athlete'
            c4.font = font(bold=True, color=WHITE, sz=8)
            c4.fill = fill(BLACK)
            c4.alignment = center()

        # ── Column headers ────────────────────────────────────────────────────
        current_row += 1
        ws.row_dimensions[current_row].height = 13
        style(ws, f'A{current_row}', 'Exercise', bold=True, fc=BLACK, bg=LGRAY, align=left(), sz=8, border=border_all(GRAY3))
        style(ws, f'B{current_row}', 'Priority', bold=True, fc=BLACK, bg=LGRAY, align=left(), sz=8, border=border_all(GRAY3))

        COACH_COLS = ['Sets', 'Reps', 'RPE', 'Range', 'Notes']
        ATH_COLS   = ['Load', 'RPE', 'Tonnage', 'Video']
        for wi in range(len(weeks)):
            wc = week_start_col(wi)
            for j, lbl in enumerate(COACH_COLS):
                style(ws, f'{get_column_letter(wc+j)}{current_row}', lbl,
                      bold=True, fc=BLACK, bg=LGRAY, align=center(), sz=7, border=border_all(GRAY3))
            for j, lbl in enumerate(ATH_COLS):
                style(ws, f'{get_column_letter(wc+5+j)}{current_row}', lbl,
                      bold=True, fc=BLACK, bg=LGRAY2, align=center(), sz=7, border=border_all(GRAY3))

        # ── Exercise rows ─────────────────────────────────────────────────────
        ex_row = current_row + 1
        for bi, block in enumerate(day_blocks):
            sets = block.get('sets', [])
            bg_ex = WHITE if bi % 2 == 0 else LGRAY2

            for si, s in enumerate(sets):
                ws.row_dimensions[ex_row].height = 14
                is_first_set = si == 0
                bg_row = WHITE if bi % 2 == 0 else LGRAY2

                # A: exercise name (only first set of each block)
                c_a = ws[f'A{ex_row}']
                c_a.value = block.get('exercise', {}).get('name', '') if is_first_set else ''
                c_a.font = font(bold=True, color=BLACK, sz=9)
                c_a.fill = fill(bg_row)
                c_a.alignment = left()
                c_a.border = border_all(GRAY3)

                # B: priority/category
                c_b = ws[f'B{ex_row}']
                c_b.value = block.get('exercise', {}).get('category', '') if is_first_set else ''
                c_b.font = font(color=BLACK, sz=8)
                c_b.fill = fill(bg_row)
                c_b.alignment = left()
                c_b.border = border_all(GRAY3)

                # Per week
                for wi, wk in enumerate(weeks):
                    wc = week_start_col(wi)
                    wk_days = wk.get('days', [])
                    if di >= len(wk_days):
                        for j in range(10):
                            cc = ws[f'{get_column_letter(wc+j)}{ex_row}']
                            cc.fill = fill(bg_row)
                            cc.border = border_all(GRAY3)
                        continue

                    wk_blocks = wk_days[di].get('blocks', [])
                    if bi >= len(wk_blocks):
                        for j in range(10):
                            cc = ws[f'{get_column_letter(wc+j)}{ex_row}']
                            cc.fill = fill(bg_row)
                            cc.border = border_all(GRAY3)
                        continue

                    wk_sets = wk_blocks[bi].get('sets', [])
                    if si >= len(wk_sets):
                        for j in range(10):
                            cc = ws[f'{get_column_letter(wc+j)}{ex_row}']
                            cc.fill = fill(bg_row)
                            cc.border = border_all(GRAY3)
                        continue

                    ws_set = wk_sets[si]
                    prev_load = resolve_load(wk_sets[si-1].get('load'), wk_sets[si-1].get('resolvedLoad')) if si > 0 else None
                    resolved = resolve_load(ws_set.get('load'), prev_load or ws_set.get('resolvedLoad'))
                    tonnage = resolved * (ws_set.get('reps') or 0)

                    def sc(j, val, **kw):
                        coord = f'{get_column_letter(wc+j)}{ex_row}'
                        cc = ws[coord]
                        cc.value = val if val else ''
                        cc.font = font(color=BLACK, sz=9, **{k: v for k, v in kw.items() if k in ['bold']})
                        cc.fill = fill(bg_row)
                        cc.alignment = center()
                        cc.border = border_all(GRAY3)

                    sc(0, 1)                              # Sets (always 1 per row)
                    sc(1, ws_set.get('reps') or '')       # Reps
                    sc(2, ws_set.get('rpe') or '')        # Coach RPE
                    sc(3, '')                             # Range
                    sc(4, block.get('notes', '') if is_first_set else '')  # Notes
                    # Athlete cols
                    cc_load = ws[f'{get_column_letter(wc+5)}{ex_row}']
                    load_val = ws_set.get('load', '')
                    display = resolved if resolved else (load_val if load_val else '')
                    cc_load.value = display if display else ''
                    cc_load.font = font(bold=True, color=BLACK, sz=9)
                    cc_load.fill = fill(LGRAY2 if bi % 2 == 0 else WHITE)
                    cc_load.alignment = center()
                    cc_load.border = border_all(GRAY3)

                    sc(6, ws_set.get('rpe') or '')        # Athlete RPE
                    sc(7, tonnage if tonnage else '')      # Tonnage
                    sc(8, '')                             # Video checkbox

                    # Override fill for athlete cols
                    for j in [6, 7, 8]:
                        ws[f'{get_column_letter(wc+j)}{ex_row}'].fill = fill(LGRAY2 if bi % 2 == 0 else WHITE)

                ex_row += 1

        # ── Gap row ───────────────────────────────────────────────────────────
        ws.row_dimensions[ex_row].height = 8
        ws.merge_cells(f'A{ex_row}:B{ex_row}')
        ws[f'A{ex_row}'].fill = fill(BLACK)
        for wi in range(len(weeks)):
            wc = week_start_col(wi)
            ws.merge_cells(f'{get_column_letter(wc)}{ex_row}:{get_column_letter(wc+9)}{ex_row}')
            ws[f'{get_column_letter(wc)}{ex_row}'].fill = fill(BLACK)

        current_row = ex_row + 2

    wb.save(sys.argv[2])
    print("OK")

if __name__ == '__main__':
    with open(sys.argv[1]) as f:
        data = json.load(f)
    build(data)
